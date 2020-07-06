#!/usr/bin/env python3

import argparse
import os
import sys

from junitparser import TestCase, TestSuite, JUnitXml
from lxml import objectify
from openpyxl import Workbook
from openpyxl.chart import Reference, BarChart3D
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

args = None
DEBUG = True

def debug(str):
    if DEBUG:
        print(str)


class TestCase:
    def __init__(self, name, refid=None, anchor=None):
        self.name = name
        self.refid = refid
        self.anchor = anchor
        self.group = None
        self.brief = ""

    def __str__(self):
        return self.name


class TestSuite:
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.testcases = []

    def get_or_create(self, name):
        for tc in self.testcases:
            if tc.name == name:
                return tc
        tc = TestCase(name=name)
        self.add(tc)
        return tc

    def get_by_id(self, refid):
        for tc in self.testcases:
            if tc.refid == refid:
                return tc
        return None

    def dump(self, group=None):
        for tc in self.testcases:
            if group:
                if group in tc.group:
                    print("{}:\n  id: {}\n  group: {}\n".format(tc.name, tc.refid, tc.group))
            else:
                print("{}:\n  id: {}\n  group: {}\n".format(tc.name, tc.refid, tc.group))

    def add(self, testcase):
        self.testcases.append(testcase)

    def stringify_children(self, node):
        parts = node.text
        for c in node.getchildren():
            parts = parts + c.text

        if node.tail:
            parts = parts + node.tail

            # filter removes possible Nones in texts and tails
            return ''.join(filter(None, parts))

    def parse(self, filename="group__all__tests.xml"):

        all_tests = objectify.parse("{}/{}".format(args.xmlroot, filename))
        groups = all_tests.xpath("//compounddef/innergroup")
        debug("====== Parsing Tests ========")
        for g in groups:

            refid = str(g.xpath('@refid')[0])
            debug(f"=> Parsing {g} with refid={refid}...")

            if args.add_test_identifiers:
                identifier = refid.replace("group__", "")
                identifier = identifier.replace("__tests", "")
                identifier = identifier.replace("__", ".")
                debug(f"Identifier: {identifier}")
            else:
                identifier = ""

            group_file = os.path.join(args.xmlroot, str(refid) + ".xml")
            debug(f"Parsing file {group_file}")

            group_tests = objectify.parse(group_file)
            title = group_tests.xpath("//compounddef/title")[0]
            testcases = group_tests.xpath("//compounddef/sectiondef/memberdef")

            for case in testcases:
                refid = case.xpath("@id")[0]

                if str(case.name).startswith(args.test_prefix):
                    brief = ""
                    brief_el = case.xpath("briefdescription/para")
                    if brief_el:
                        b = brief_el[0]
                        brief = b.text
                        if isinstance(b, objectify.ObjectifiedElement) and b.getchildren():
                            for i in b.getchildren():
                                brief = brief + i.text

                            if b.tail:
                                brief = brief + b.tail

                    if args.add_test_identifiers:
                        n = str(case.name).replace("test_", identifier + ".")
                    else:
                        n = str(case.name)

                    tc = TestCase(n, refid=str(refid))
                    tc.group = str(title)
                    tc.brief = brief
                    self.add(tc)

        debug("====== Done Parsing Tests ========")


class Implementation:
    def __init__(self, name, refid=None, anchor=None, file=None, line=0):
        self.name = name
        self.refid = refid
        self.anchor = anchor
        self.file = file
        self.line = line


class Requirement:
    def __init__(self, name, refid=None, anchor=None):
        self.name = name
        self.refid = refid
        self.anchor = anchor
        self.tests = []
        self.implementations = []
        self.title = None
        self.text = None
        self.parent = None

    def add_test(self, test):
        self.tests.append(test)

    def add_implementation(self, implementation):
        self.implementations.append(implementation)

    def dump(self):
        print(f"Name={self.name} ({self.refid})")
        print(f"  Title: {self.title}")
        print(f"  Text: {self.text}")
        print(f"  Parent: {self.parent}")
        print("  Tests:")
        for t in self.tests:
            print(f"   - {t.name}")
        print("  Implementations:")
        for i in self.implementations:
            print("   - function: {}".format(i.name))
            print("   - file: {}".format(i.file))
            print("   - line: {}".format(i.line))
        print("\n")


class Requirements:
    def __init__(self):
        self.requirements = []

    def add(self, req):
        self.requirements.append(req)

    def get_or_create(self, name):
        for r in self.requirements:
            if r.name == name:
                return r
        r = Requirement(name=name)
        self.add(r)
        return r

    def tests(self):
        tests = set()
        for r in self.requirements:
            tests.update(r.tests)
        return tests

    def verifies(self, req):
        return self.get_or_create(req).tests

    def satisfies(self, req):
        return self.get_or_create(req).implementations

    def requirement(self, test):
        reqs = set()
        for r in self.requirements:
            for t in r.tests:
                if t.name == test:
                    reqs.add(r.name)
        return reqs

    def dump(self):
        for r in self.requirements:
            r.dump()

    def get_by_parent(self, parent):
        ret = []
        for r in self.requirements:
            if r.parent == parent:
                ret.append(r)
        return ret

    def get_by_test(self, testcase):
        ret = []
        for r in self.requirements:
            for t in r.tests:
                if t.name == testcase:
                    ret.append(r.name)
        return ret

    def _add_req(self, req_file, req, level, parent):
        xml_id = req.attrib['id']
        reqid = str(xml_id).replace("Requirements_1", "")
        debug(f"requirement: {reqid}")
        r = self.get_or_create(reqid)
        r.name = reqid
        r.refid = xml_id
        req_details_list = req_file.xpath(f"//sect{level}[@id='{r.refid}']")

        text = None
        if req_details_list:
            req_details = req_details_list[0]
            title = str(req_details.xpath("title")[0])
            r.title = title.replace(f"{reqid}:", "")
            text = req_details.xpath("para")
            if text:
                r.text = str(text[0])
        r.parent = parent

        return reqid

    def parse(self, filename="Requirements.xml"):

        file_path = os.path.join(args.xmlroot, filename)
        if not os.path.exists(file_path):
            sys.exit(f"{file_path} does not exist...")

        req_file = objectify.parse(file_path)
        req_list = req_file.xpath("//sect1")

        for req in req_list:
            name = self._add_req(req_file, req, 1, None)
            for subreq in req.xpath("//sect2"):
                self._add_req(req_file, subreq, 2, name)


class TestReport:

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.testcases = []
        self.passed = 0
        self.failed = 0
        self.errors = 0
        self.skipped = 0

    def parse(self, file):
        junit_xml = JUnitXml.fromfile(file)
        for suite in junit_xml:
            self.passed += suite.tests - suite.errors - suite.failures - suite.skipped
            self.failed += suite.failures
            self.skipped += suite.skipped
            self.errors += suite.errors

            for testcase in suite:
                self.testcases.append(testcase)


class RTM():
    def __init__(self, xml_root, report_file=None):
        self.rtm = {}
        self.rtm_req = []
        self.req_id_list = []
        self.xml_root = xml_root
        self.report_file = report_file

        self.requirements = None
        self.suite = None

        self.groups = dict()

    def find_file(self, group_dict, struct_dict, refid):
        for k, refs in group_dict.items():
            if refid in refs:
                return k
        for k, refs in struct_dict.items():
            if refid in refs:
                return k

    def parse_index(self):
        index_file = os.path.join(self.xml_root, "index.xml")
        index = objectify.parse(index_file)
        groups = index.xpath('//compound[@kind="group"]')
        for group in groups:
            refid = group.attrib['refid']
            name = group.name
            self.groups[refid] = name

    def parse_xml(self):
        self.parse_index()

        try:
            verify_file = os.path.join(self.xml_root, "verify.xml")
            verify_list = objectify.parse(verify_file)
            varlist = verify_list.xpath("//variablelist/node()")

            for ve in varlist:
                refs = ve.xpath("term/ref")
                for ref in refs:

                    refid = str(ref.xpath("@refid")[0])
                    debug(f"Working on {ref} ({refid})...")

                    debug(f"Check if we have {refid} already...")
                    found = self.suite.get_by_id(refid)
                    if found:
                        debug("found")
                        tc = found
                    else:
                        tc = self.suite.get_or_create(str(ref))
                        tc.refid = refid

                    debug(tc)

                reqs = ve.xpath("para/ref")
                for req in reqs:
                    r = self.requirements.get_or_create(str(req))
                    r.add_test(tc)
        except OSError as e:
            print(str(e))
            sys.exit(2)

        try:
            satisfy_list = objectify.parse("{}/satisfy.xml".format(self.xml_root))
            varlist = satisfy_list.xpath("//variablelist/node()")
            for ve in varlist:
                refs = ve.xpath("term/ref")
                for ref in refs:
                    refid = str(ref.xpath("@refid")[0])
                    debug(refid)
                    group = refid.split("_ga", 1)[0]
                    debug(group)
                    xmlid = refid.split("_1ga", 1)[-1]
                    debug(xmlid)
                    group_file = refid.split("_1ga", 1)[0]
                    if xmlid not in group:
                        func_id = "{}_1ga{}".format(group, xmlid)
                    else:
                        func_id = group

                    if os.path.exists(os.path.join(self.xml_root, f"{group_file}.xml")):
                        xmlfile = "{}.xml".format(group_file)
                    else:
                        xmlfile = "{}.xml".format(group)
                    debug(xmlfile)
                    debug(func_id)
                    impl = objectify.parse(os.path.join(self.xml_root, xmlfile))

                    file_list = impl.xpath("//*[@id='{}']//location/@file".format(func_id))
                    if file_list:
                        file = file_list[0]
                        line = impl.xpath("//*[@id='{}']//location/@line".format(func_id))[0]

                    impl = Implementation(
                        name=str(ref),
                        refid=refid,
                        file=file,
                        line=line
                    )

                listitems = ve.xpath("para/ref")
                for item in listitems:
                    requirement = str(item)
                    r = self.requirements.get_or_create(requirement)
                    r.add_implementation(impl)

        except OSError as e:
            print(str(e))
            sys.exit(2)

    @staticmethod
    def adjust_width(worksheet):
        wrap = False
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                if cell.alignment.wrapText:
                    wrap = True
                if cell.coordinate in worksheet.merged_cells:  # not check merge_cells
                    continue
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            if wrap:
                wrap = False
                continue
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width

    def write_srs(self, filename):
        SRS_HEADER = """
=====================
Software Requirements
=====================

This is the Software Requirements Specification.


Software capabilities
=====================
"""
        with open(filename, "w") as req:
            req.write(SRS_HEADER)
            for r in self.requirements.requirements:
                if not r.parent:
                    dashes = len(r.title) * "-"
                    req.write(f"\n{r.title}\n{dashes}\n\n")

                    for c in self.requirements.get_by_parent(r.name):
                        req.write(f".. item:: {c.name} {c.title}\n")
                        # req.write("\n")
                        for t in c.tests:
                            req.write(f"    :validated_by: {t.name}\n")
                        req.write(f"\n    {c.text.rstrip()}\n\n")

    def write_rst(self, output_dir="docs"):

        out = os.path.join(output_dir, "SRS.rst")
        self.write_srs(out)

        print(f"Number of testcases: {len(self.suite.testcases)}")
        test_file = os.path.join(output_dir, "tests.rst")
        with open(test_file, "w") as fp:
            section = "Integration Tests"
            fp.write(f"{section}\n{len(section) * '='}\n\n")
            for item in self.suite.testcases:
                fp.write(f".. item:: {item.name} {item.brief}\n")
                # for v in self.requirements.get_by_test(item.name):
                #    fp.write(f"    :validates: {v}\n")
                fp.write("\n")

    def write_xls(self, output_file):
        book = Workbook()
        sheet1 = book.active
        sheet1.title = "RTM"
        testplan_sheet = book.create_sheet("Tests")
        sheet3 = book.create_sheet("Requirements")

        # sheet1.row(0).height_mismatch = True
        sheet1.row_dimensions[1].height = 70

        # Write Testplan
        testplan_sheet.column_dimensions['A'].width = 50
        testplan_sheet.column_dimensions['B'].width = 50
        testplan_sheet.column_dimensions['C'].width = 50

        cell = testplan_sheet.cell(1, 1, "ID")
        cell.font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='top')

        cell = testplan_sheet.cell(1, 2, "Brief")
        cell.alignment = header_alignment
        cell.font = Font(bold=True)

        cell = testplan_sheet.cell(1, 3, "Details")
        cell.alignment = header_alignment
        cell.font = Font(bold=True)

        row = 2
        test_links = {}
        for tc in self.suite.testcases:
            testplan_sheet.cell(row, 1, tc.name)
            if tc.brief != "":
                testplan_sheet.cell(row, 2, tc.brief)
            test_links[tc.name] = 'A{}'.format(row)
            row = row + 1

        # Write Requirements sheet
        sheet3.column_dimensions['A'].width = 50
        sheet3.column_dimensions['B'].width = 50
        sheet3.column_dimensions['C'].width = 200

        wrap = Alignment(vertical='top', wrap_text=True)
        cell = sheet3.cell(1, 1, "ID")
        cell.alignment = header_alignment
        cell.font = Font(bold=True)

        cell = sheet3.cell(1, 2, "Title")
        cell.alignment = header_alignment
        cell.font = Font(bold=True)

        cell = sheet3.cell(1, 3, "text")
        cell.alignment = header_alignment
        cell.font = Font(bold=True)

        row = 2
        links = {}
        for r in self.requirements.requirements:
            sheet3.cell(row, 1, r.name)
            sheet3.cell(row, 2, r.title)
            cell = sheet3.cell(row, 3, r.text)
            cell.alignment = wrap
            links[r.name] = 'A{}'.format(row)
            row = row + 1

        row = 3
        req_cols = {}
        # Now go through all tests
        for t in self.suite.testcases:
            cell = sheet1.cell(row, 1, '=HYPERLINK("#Tests!{}","{}")'.format(test_links[t.name], t.name))
            # cell.font = Font(bold=True)
            # header_alignment = Alignment(horizontal='center', vertical='top')
            # cell.alignment = header_alignment

            sheet1.column_dimensions['A'].width = 50
            # Get all requirements for a test
            for r in self.requirements.requirement(t.name):
                if r not in req_cols:
                    debug(f"{1}x{len(req_cols) + 2}")
                    cell = sheet1.cell(1, len(req_cols) + 3, '=HYPERLINK("#Requirements!{}","{}")'.format(links[r], r))
                    header_alignment1 = Alignment(horizontal='center', vertical='top', textRotation=90,
                                                  shrinkToFit=True)
                    cell.font = Font(bold=True)
                    cell.alignment = header_alignment1

                    col = get_column_letter(cell.column)
                    sheet1.column_dimensions[col].width = 3

                    # sheet1.col(len(req_cols) + 2).width = 256 * 4
                    req_cols[r] = len(req_cols) + 3

                sheet1.cell(row, req_cols[r], "X")

            row += 1

        for r in self.requirements.requirements:
            if r.name not in req_cols:
                debug(f"{1}x{len(req_cols) + 2}")
                cell = sheet1.cell(1, len(req_cols) + 3,
                                   '=HYPERLINK("#Requirements!{}","{}")'.format(links[r.name], r.name))
                header_alignment = Alignment(horizontal='center', vertical='top', textRotation=90, shrinkToFit=True)
                cell.alignment = header_alignment
                cell.font = Font(bold=True)

                col = get_column_letter(cell.column)
                sheet1.column_dimensions[col].width = 3

                req_cols[r.name] = len(req_cols) + 2

        sheet1.cell(1, 1, "Requirements")
        sheet1.cell(2, 1, "Test Cases")
        sheet1.cell(1, 2, "Reqs Tested")

        row1 = 3
        tests = self.suite.testcases
        for t in tests:
            cell = "=COUNTA({}{}:{}{})".format(get_column_letter(3), row1,
                                               get_column_letter(1 + len(self.requirements.requirements)), row1)
            sheet1.cell(row1, 2, cell)
            row1 = row1 + 1

        sheet1.cell(2, 2, "=SUM(B3:B{})".format(3 + len(tests)))

        col = 3
        for t in self.requirements.requirements:
            sheet1.cell(2, col,
                        "=COUNTA({}3:{}{})".format(get_column_letter(col), get_column_letter(col), 3 + len(tests)))
            col = col + 1

        # Report
        results = {}
        if self.report_file and os.path.exists(self.report_file):
            test_report = book.create_sheet("Test Report")
            cell = test_report.cell(11, 1, "Testcase")
            test_report.column_dimensions['A'].width = 60

            cell.font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='top')
            cell.alignment = header_alignment

            r = TestReport()
            r.parse(self.report_file)

            tidx = 2
            targets = []
            tt = {}
            for t in r.testcases:
                target, test = t.classname.split(":")
                res = "pass"
                if t.result:
                    res = t.result.type
                if not tt.get(t.name, None):
                    tt[t.name] = {target: res}
                else:
                    tt[t.name][target] = res

                # Targets (Header)
                if target not in targets:
                    for c in [1, 11]:
                        cell = test_report.cell(c, tidx, target)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='top', textRotation=90,
                                                   shrinkToFit=True)
                    targets.append(target)
                    test_report.column_dimensions[get_column_letter(tidx)].width = 8
                    tidx += 1

            summary = ["Pass", "Fail", "Skipped", "Error", "Total", "Pass Rate", "Execution Rate"]

            row = 2
            for item in summary:
                cell = test_report.cell(row, 1, item)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='right', vertical='top')
                row += 1

            row = 12
            for t in tt.keys():
                test_report.cell(row, 1, t)
                for target in tt[t].keys():
                    test_report.cell(row, targets.index(target) + 2, tt[t][target])

                results[t] = 'A{}'.format(row + 1)
                row = row + 1

            row = 2
            summary = ["pass", "failure", "skipped", "error", "total", "passrate", "execution"]
            for item in summary:
                col = 2
                for t in targets:
                    style = 'Neutral'
                    if item == "total":
                        formula = '=SUM({col}$2:{col}$5)'.format(col=get_column_letter(col))
                        style = "Total"
                    elif item == "passrate":
                        formula = '={col}2/({col}6-{col}4)'.format(col=get_column_letter(col))
                        style = "Percent"
                    elif item == "execution":
                        formula = "=({col}2+{col}3+{col}5)/{col}6".format(col=get_column_letter(col))
                        style = "Percent"
                    elif item == "skipped":
                        formula = '=COUNTIF({}$12:{}${},"{}") + COUNTBLANK({}$12:{}${})'.format(get_column_letter(col),
                                                                                                get_column_letter(col),
                                                                                                len(tt), item,
                                                                                                get_column_letter(col),
                                                                                                get_column_letter(col),
                                                                                                len(tt))
                    else:
                        formula = '=COUNTIF({}$12:{}${},"{}")'.format(get_column_letter(col), get_column_letter(col),
                                                                      len(tt), item)

                    cell = test_report.cell(row, col, formula)
                    cell.style = style
                    col += 1
                row += 1

            chart1 = BarChart3D()
            chart1.style = 11
            chart1.grouping = "percentStacked"
            chart1.overlap = 100
            chart1.title = 'Platforms'

            data = Reference(test_report, min_col=1, min_row=2, max_row=5, max_col=len(targets) + 1)
            cats = Reference(test_report, min_col=2, min_row=1, max_col=len(targets) + 1)

            chart1.add_data(data, from_rows=True, titles_from_data=True, )
            chart1.set_categories(cats)
            chart1.width = 25
            chart1.height = 15

            # chart1.shape = 4
            test_report.add_chart(chart1, "{}10".format(get_column_letter(len(targets) + 5)))

        self.adjust_width(testplan_sheet)
        self.adjust_width(sheet3)

        book.save(output_file)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Generate Requirement Traceability Matrix (RTM).")

    parser.add_argument('-x', '--xmlroot', default=None,
                        help="Root directory of XML files generated by doxygen.")

    parser.add_argument('-r', '--rtm-file',
                        help="RTM file in Excel format")

    parser.add_argument('--generate-rst', action="store_true",
                        help="Generate RST files")

    parser.add_argument('--output-dir',
                        help="Output directory")

    parser.add_argument('-g', '--group',
                        help="filter by group")

    parser.add_argument('-p', '--test-prefix', default="test_",
                        help="Testcase prefix to identify testcases")

    parser.add_argument('-i', '--add-test-identifiers', action="store_true",
                        help="Add additional identifiers based on test groups")

    parser.add_argument('-d', '--dump', action="store_true",
                        help="Dump requirements and dependencies")

    parser.add_argument("-j", "--junit-file", help="file with test results in junit format")
    return parser.parse_args()


def main():
    global args

    args = parse_args()
    if not args.xmlroot:
        sys.exit("No XML root given")

    if args.junit_file and not os.path.exists(args.junit_file):
        sys.exit("File {} does not exist.".format(args.junit_file))

    rtm = RTM(xml_root=args.xmlroot, report_file=args.junit_file)

    rtm.suite = TestSuite()
    rtm.suite.parse()
    if args.group:
        rtm.suite.dump(args.group)
        sys.exit(0)

    rtm.requirements = Requirements()
    rtm.requirements.parse()

    rtm.parse_xml()

    if args.dump:
        rtm.requirements.dump()
        sys.exit(0)

    if args.rtm_file:
        rtm.write_xls(args.rtm_file)

    if args.generate_rst:
        print("Generating RST files...")
        rtm.write_rst(args.output_dir)


if __name__ == "__main__":
    main()
